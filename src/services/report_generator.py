"""Report Generation Service for creating Excel/CSV reports."""

import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Service for generating formatted reports in Excel and CSV formats."""

    # Syatt Design System Colors
    PURPLE_PRIMARY = "554DFF"
    PURPLE_GRADIENT = "7D00FF"
    MINT_PRIMARY = "00FFCE"
    GRAY_LIGHT = "F5F5F5"
    GRAY_MED = "CCCCCC"
    WHITE = "FFFFFF"
    RED = "FF0000"
    ORANGE = "FFA500"
    GREEN = "00FF00"

    def __init__(self):
        """Initialize the report generator."""
        pass

    def generate_epic_reconciliation_report(
        self,
        month: str,
        epic_data: List[Dict[str, Any]],
        project_summary: Dict[str, Any],
    ) -> BytesIO:
        """
        Generate monthly epic reconciliation report in Excel format.

        Args:
            month: Month in YYYY-MM format
            epic_data: List of epic hour records with forecast/actual
            project_summary: Summary data for executive overview

        Returns:
            BytesIO object containing Excel file
        """
        logger.info(f"Generating epic reconciliation report for {month}")

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Executive Summary
        self._create_executive_summary_sheet(wb, month, project_summary)

        # Sheet 2: Epic Details
        self._create_epic_details_sheet(wb, month, epic_data)

        # Sheet 3: Variance Analysis (only >10% variance)
        high_variance_epics = [
            e for e in epic_data if abs(e.get("variance_pct", 0)) > 10
        ]
        self._create_variance_analysis_sheet(wb, month, high_variance_epics)

        # Sheet 4: Team Performance
        self._create_team_performance_sheet(wb, month, epic_data)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Generated report with {len(epic_data)} epics")
        return output

    def _create_executive_summary_sheet(
        self, wb: openpyxl.Workbook, month: str, summary: Dict[str, Any]
    ) -> None:
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary", 0)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = (
            f"Monthly Epic Reconciliation - {datetime.strptime(month, '%Y-%m').strftime('%B %Y')}"
        )
        ws["A1"].font = Font(size=16, bold=True, color=self.PURPLE_PRIMARY)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Generated date
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws["A2"].font = Font(size=10, italic=True)

        # Spacer
        row = 4

        # Key metrics table
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col)
            cell.value = header
            cell.font = Font(bold=True, color=self.WHITE)
            cell.fill = PatternFill(
                start_color=self.PURPLE_PRIMARY,
                end_color=self.PURPLE_PRIMARY,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Metrics data
        metrics = [
            ("Total Projects", summary.get("total_projects", 0)),
            ("Total Epics Analyzed", summary.get("total_epics", 0)),
            ("Forecast Hours (Total)", f"{summary.get('total_forecast_hours', 0):.1f}"),
            ("Actual Hours (Total)", f"{summary.get('total_actual_hours', 0):.1f}"),
            ("Overall Variance", f"{summary.get('total_variance_pct', 0):.1f}%"),
            ("Epics Over Budget (>10%)", summary.get("epics_over_budget", 0)),
            ("Epics Under Budget (<-10%)", summary.get("epics_under_budget", 0)),
        ]

        for metric_name, metric_value in metrics:
            ws.cell(row, 1).value = metric_name
            ws.cell(row, 2).value = metric_value

            # Color code variance
            if "Variance" in metric_name:
                variance = summary.get("total_variance_pct", 0)
                if abs(variance) > 20:
                    ws.cell(row, 2).font = Font(bold=True, color=self.RED)
                elif abs(variance) > 10:
                    ws.cell(row, 2).font = Font(bold=True, color=self.ORANGE)

            row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_epic_details_sheet(
        self, wb: openpyxl.Workbook, month: str, epic_data: List[Dict[str, Any]]
    ) -> None:
        """Create detailed epic breakdown sheet."""
        ws = wb.create_sheet("Epic Details")

        # Headers
        headers = [
            "Project",
            "Epic Key",
            "Epic Name",
            "Team",
            "Forecast Hours",
            "Actual Hours",
            "Variance Hours",
            "Variance %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color=self.WHITE)
            cell.fill = PatternFill(
                start_color=self.PURPLE_PRIMARY,
                end_color=self.PURPLE_PRIMARY,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        row = 2
        for epic in epic_data:
            ws.cell(row, 1).value = epic.get("project_key", "")
            ws.cell(row, 2).value = epic.get("epic_key", "")
            ws.cell(row, 3).value = epic.get("epic_name", "Unknown")
            ws.cell(row, 4).value = epic.get("team", "")
            ws.cell(row, 5).value = epic.get("forecast_hours", 0)
            ws.cell(row, 5).number_format = "0.0"
            ws.cell(row, 6).value = epic.get("actual_hours", 0)
            ws.cell(row, 6).number_format = "0.0"
            ws.cell(row, 7).value = epic.get("variance_hours", 0)
            ws.cell(row, 7).number_format = "0.0"
            ws.cell(row, 8).value = epic.get("variance_pct", 0)
            ws.cell(row, 8).number_format = "0.0%"

            # Color code variance percentage
            variance_pct = epic.get("variance_pct", 0)
            if abs(variance_pct) > 20:
                ws.cell(row, 8).font = Font(bold=True, color=self.RED)
            elif abs(variance_pct) > 10:
                ws.cell(row, 8).font = Font(bold=True, color=self.ORANGE)

            row += 1

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        ws.column_dimensions["C"].width = 40  # Epic name

    def _create_variance_analysis_sheet(
        self,
        wb: openpyxl.Workbook,
        month: str,
        high_variance_epics: List[Dict[str, Any]],
    ) -> None:
        """Create variance analysis sheet (>10% variance only)."""
        ws = wb.create_sheet("Variance Analysis")

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "High Variance Epics (>10%)"
        ws["A1"].font = Font(size=14, bold=True, color=self.RED)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "Project",
            "Epic",
            "Team",
            "Forecast",
            "Actual",
            "Variance",
            "Variance %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(3, col)
            cell.value = header
            cell.font = Font(bold=True, color=self.WHITE)
            cell.fill = PatternFill(
                start_color=self.PURPLE_PRIMARY,
                end_color=self.PURPLE_PRIMARY,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sort by abs(variance_pct) descending
        sorted_epics = sorted(
            high_variance_epics,
            key=lambda x: abs(x.get("variance_pct", 0)),
            reverse=True,
        )

        # Data rows
        row = 4
        for epic in sorted_epics:
            ws.cell(row, 1).value = epic.get("project_key", "")
            ws.cell(row, 2).value = (
                f"{epic.get('epic_key', '')} - {epic.get('epic_name', '')}"
            )
            ws.cell(row, 3).value = epic.get("team", "")
            ws.cell(row, 4).value = epic.get("forecast_hours", 0)
            ws.cell(row, 4).number_format = "0.0"
            ws.cell(row, 5).value = epic.get("actual_hours", 0)
            ws.cell(row, 5).number_format = "0.0"
            ws.cell(row, 6).value = epic.get("variance_hours", 0)
            ws.cell(row, 6).number_format = "0.0"
            ws.cell(row, 7).value = epic.get("variance_pct", 0)
            ws.cell(row, 7).number_format = "0.0%"

            # Highlight over-budget in red, under-budget in orange
            variance_pct = epic.get("variance_pct", 0)
            if variance_pct > 10:
                ws.cell(row, 7).font = Font(bold=True, color=self.RED)
            else:
                ws.cell(row, 7).font = Font(bold=True, color=self.ORANGE)

            row += 1

        # Auto-size columns
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        for col in range(4, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_team_performance_sheet(
        self, wb: openpyxl.Workbook, month: str, epic_data: List[Dict[str, Any]]
    ) -> None:
        """Create team performance summary sheet."""
        ws = wb.create_sheet("Team Performance")

        # Aggregate by team
        from collections import defaultdict

        team_stats = defaultdict(lambda: {"forecast": 0.0, "actual": 0.0, "count": 0})

        for epic in epic_data:
            team = epic.get("team", "Unassigned")
            team_stats[team]["forecast"] += epic.get("forecast_hours", 0)
            team_stats[team]["actual"] += epic.get("actual_hours", 0)
            team_stats[team]["count"] += 1

        # Headers
        headers = [
            "Team",
            "# Epics",
            "Forecast Hours",
            "Actual Hours",
            "Variance Hours",
            "Variance %",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color=self.WHITE)
            cell.fill = PatternFill(
                start_color=self.PURPLE_PRIMARY,
                end_color=self.PURPLE_PRIMARY,
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        row = 2
        for team, stats in sorted(team_stats.items()):
            forecast = stats["forecast"]
            actual = stats["actual"]
            variance = actual - forecast
            variance_pct = (variance / forecast * 100) if forecast > 0 else 0

            ws.cell(row, 1).value = team
            ws.cell(row, 2).value = stats["count"]
            ws.cell(row, 3).value = forecast
            ws.cell(row, 3).number_format = "0.0"
            ws.cell(row, 4).value = actual
            ws.cell(row, 4).number_format = "0.0"
            ws.cell(row, 5).value = variance
            ws.cell(row, 5).number_format = "0.0"
            ws.cell(row, 6).value = (
                variance_pct / 100
            )  # Convert to decimal for percentage format
            ws.cell(row, 6).number_format = "0.0%"

            # Color code variance
            if abs(variance_pct) > 20:
                ws.cell(row, 6).font = Font(bold=True, color=self.RED)
            elif abs(variance_pct) > 10:
                ws.cell(row, 6).font = Font(bold=True, color=self.ORANGE)

            row += 1

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
